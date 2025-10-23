"""
질문 로거 - 답변 못 찾은 질문들을 엑셀에 기록
엑셀 파일을 열어놓은 상태에서도 작동 가능

핵심 기능:
    1. 미답변 질문 자동 기록
    2. 중복 질문 자동 제거
    3. 엑셀 열린 상태 대응 (재시도 + 임시 파일)
    4. 자동 병합

작동 방식:
    1. 정상: 바로 엑셀에 저장
    2. 엑셀 열림: 5번 재시도 → 실패 시 임시 파일에 저장
    3. 엑셀 닫음: 로그 페이지 접속 시 자동 병합

파일:
    - logs/unanswered_questions.xlsx (메인 로그)
    - logs/unanswered_questions_temp.txt (임시 저장)
"""
import pandas as pd
import os
from datetime import datetime
import logging
import time
from openpyxl import load_workbook, Workbook

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class QuestionLogger:
    """
    알 수 없는 질문 로깅 클래스
    
    Attributes:
        log_file: 메인 로그 파일 경로
        temp_log_file: 임시 로그 파일 경로
    """
    
    def __init__(self, log_file='logs/unanswered_questions.xlsx'):
        """
        질문 로거 초기화
        
        Args:
            log_file: 로그 파일 경로 (기본값: logs/unanswered_questions.xlsx)
        
        Process:
            1. 로그 파일 경로 설정
            2. 임시 파일 경로 설정 (_temp.txt)
            3. 로그 파일 존재 확인 (없으면 생성)
        """
        self.log_file = log_file
        self.temp_log_file = log_file.replace('.xlsx', '_temp.txt')
        self._ensure_log_file_exists()
    
    def _ensure_log_file_exists(self):
        """
        로그 파일이 없으면 생성
        
        Process:
            1. logs 폴더가 없으면 생성
            2. 엑셀 파일이 없으면 빈 파일 생성
            3. 컬럼: [번호, 질문, 일시, 상태, 비고]
        """
        # logs 폴더 생성
        log_dir = os.path.dirname(self.log_file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir)
            logger.info(f"로그 폴더 생성: {log_dir}")
        
        # 파일이 없으면 새로 생성
        if not os.path.exists(self.log_file):
            # 빈 DataFrame 생성 (컬럼만)
            df = pd.DataFrame(columns=[
                '번호',      # 자동 증가 번호
                '질문',      # 사용자가 입력한 질문
                '일시',      # 질문한 시간
                '상태',      # 미답변 / 답변완료
                '비고'       # 관리자 메모
            ])
            df.to_excel(self.log_file, index=False)
            logger.info(f"로그 파일 생성: {self.log_file}")
    
    def log_unknown_question(self, question: str, max_retries=5):
        """
        알 수 없는 질문 기록 (엑셀 파일을 열어놓은 상태에서도 작동)
        
        Args:
            question: 기록할 질문
            max_retries: 최대 재시도 횟수 (기본값: 5번)
        
        Process:
            1. 엑셀 파일 열기 시도
            2. 중복 체크 (같은 질문 이미 있으면 skip)
            3. 새 행 추가
            4. 저장 시도
            
            실패 시:
            - PermissionError (파일 잠김)
              → 5번 재시도 (각 1초 대기)
              → 최종 실패 시 임시 파일에 저장
        
        Example:
            성공: "WiFi 비밀번호" → logs/unanswered_questions.xlsx
            실패: 엑셀 열림 → logs/unanswered_questions_temp.txt
        
        Note:
            - openpyxl 사용 (pandas보다 안정적)
            - 중복 자동 제거
            - 임시 파일은 나중에 자동 병합
        """
        # 최대 5번 재시도
        for attempt in range(max_retries):
            try:
                # openpyxl로 파일 열기 (pandas보다 안정적)
                if os.path.exists(self.log_file):
                    # 기존 파일 열기
                    wb = load_workbook(self.log_file)
                    ws = wb.active
                    
                    # 이미 같은 질문이 있는지 확인 (중복 방지)
                    existing_questions = []
                    for row in ws.iter_rows(min_row=2, values_only=True):  # 2행부터 (헤더 제외)
                        if row[1]:  # 질문 컬럼 (2번째, 0-based index 1)
                            existing_questions.append(row[1])
                    
                    # 중복이면 skip
                    if question in existing_questions:
                        logger.info(f"이미 로그된 질문: {question}")
                        wb.close()
                        return
                    
                    # 새 행 추가
                    next_row = ws.max_row + 1  # 다음 행 번호
                    new_data = [
                        next_row - 1,  # 번호 (헤더 제외한 순번)
                        question,      # 질문
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # 일시
                        '미답변',      # 상태
                        ''            # 비고 (빈 문자열)
                    ]
                    ws.append(new_data)
                else:
                    # 파일이 없으면 새로 생성
                    wb = Workbook()
                    ws = wb.active
                    # 헤더 추가
                    ws.append(['번호', '질문', '일시', '상태', '비고'])
                    # 첫 데이터 추가
                    ws.append([1, question, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '미답변', ''])
                
                # 저장 시도
                wb.save(self.log_file)
                wb.close()
                logger.info(f"질문 로그 추가 성공: {question}")
                return  # 성공하면 함수 종료
                
            except PermissionError:
                # 파일이 열려있어서 저장 못함
                if attempt < max_retries - 1:
                    # 아직 재시도 가능 → 1초 대기 후 재시도
                    logger.warning(f"엑셀 파일이 열려있습니다. {attempt + 1}초 후 재시도... ({attempt + 1}/{max_retries})")
                    time.sleep(1)
                else:
                    # 최대 재시도 초과 → 임시 파일에 저장
                    logger.warning(f"엑셀 파일에 저장 실패. 임시 파일에 저장합니다: {self.temp_log_file}")
                    self._save_to_temp_file(question)
                    logger.info("TIP: 엑셀 파일을 닫으면 임시 파일 내용이 자동으로 병합됩니다.")
            
            except Exception as e:
                # 기타 오류
                logger.error(f"로그 저장 실패: {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    break
    
    def _save_to_temp_file(self, question: str):
        """
        임시 파일에 질문 저장 (엑셀이 열려있을 때 사용)
        
        Args:
            question: 저장할 질문
        
        Format:
            타임스탬프|질문
            예: 2025-09-30 14:30:15|WiFi 비밀번호는?
        
        Note:
            - 텍스트 파일이므로 엑셀 잠김 문제 없음
            - 나중에 엑셀 닫으면 자동 병합
        """
        try:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            # append 모드로 파일 열기 (기존 내용 유지)
            with open(self.temp_log_file, 'a', encoding='utf-8') as f:
                f.write(f"{timestamp}|{question}\n")
            logger.info(f"임시 파일에 저장 완료: {question}")
        except Exception as e:
            logger.error(f"임시 파일 저장 실패: {str(e)}")
    
    def merge_temp_file(self):
        """
        임시 파일의 내용을 엑셀에 병합 (자동 호출)
        
        Process:
            1. 임시 파일 존재 확인
            2. 임시 파일 읽기 (타임스탬프|질문 형식)
            3. 엑셀 파일 열기
            4. 각 질문을 엑셀에 추가
            5. 임시 파일 삭제
        
        When:
            - /api/unanswered 호출 시마다 실행
            - 로그 페이지 접속 시
            - 로그 페이지 자동 새로고침 시 (5초마다)
        
        Note:
            - 임시 파일이 없으면 아무것도 안 함
            - 관리자가 엑셀 닫은 후 자동 병합
        """
        # 임시 파일이 없으면 skip
        if not os.path.exists(self.temp_log_file):
            return
        
        try:
            # 1. 임시 파일 읽기
            with open(self.temp_log_file, 'r', encoding='utf-8') as f:
                temp_questions = f.readlines()
            
            # 빈 파일이면 skip
            if not temp_questions:
                return
            
            # 2. 엑셀 파일 열기
            wb = load_workbook(self.log_file)
            ws = wb.active
            
            # 3. 임시 파일의 모든 질문 추가
            for line in temp_questions:
                # 파싱: "2025-09-30 14:30:15|WiFi 비밀번호?"
                timestamp, question = line.strip().split('|', 1)
                next_row = ws.max_row + 1
                ws.append([next_row - 1, question, timestamp, '미답변', ''])
            
            # 4. 저장
            wb.save(self.log_file)
            wb.close()
            
            # 5. 임시 파일 삭제 (병합 완료)
            os.remove(self.temp_log_file)
            logger.info(f"임시 파일 병합 완료: {len(temp_questions)}개 질문")
            
        except Exception as e:
            logger.error(f"임시 파일 병합 실패: {str(e)}")
    
    def get_unanswered_count(self) -> int:
        """
        미답변 질문 개수 조회
        
        Returns:
            int: 미답변 질문 개수
        
        Used by:
            - 웹 UI (배지 표시)
            - 로그 페이지 (통계)
        """
        try:
            df = pd.read_excel(self.log_file)
            # '상태' 컬럼이 '미답변'인 행 개수
            return len(df[df['상태'] == '미답변'])
        except:
            return 0
    
    def get_all_questions(self):
        """
        모든 로그된 질문 조회
        
        Returns:
            DataFrame: 전체 로그 데이터
            컬럼: [번호, 질문, 일시, 상태, 비고]
        
        Used by:
            - /api/unanswered (로그 조회 API)
            - 로그 페이지 표시
        """
        try:
            df = pd.read_excel(self.log_file)
            return df
        except:
            # 파일 없거나 오류 시 빈 DataFrame
            return pd.DataFrame()
    
    def mark_as_answered(self, question: str, note: str = ''):
        """
        질문을 답변 완료로 표시 (수동 관리용)
        
        Args:
            question: 답변 완료 처리할 질문
            note: 비고 메모 (선택사항)
        
        Process:
            1. 엑셀 파일 읽기
            2. 해당 질문 찾기
            3. 상태를 '답변완료'로 변경
            4. 비고 업데이트 (있으면)
            5. 저장
        
        Example:
            mark_as_answered("WiFi 비밀번호?", "답변 추가함")
        """
        try:
            df = pd.read_excel(self.log_file)
            
            # 해당 질문 찾기 (마스크 생성)
            mask = df['질문'] == question
            if mask.any():
                # 상태 변경
                df.loc[mask, '상태'] = '답변완료'
                # 비고 추가 (있으면)
                if note:
                    df.loc[mask, '비고'] = note
                
                # 저장
                df.to_excel(self.log_file, index=False)
                logger.info(f"질문 답변 완료로 표시: {question}")
            
        except Exception as e:
            logger.error(f"상태 업데이트 실패: {str(e)}")
